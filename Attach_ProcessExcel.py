# Attach_ProcessExcel.py
import sys
import time
import os
import traceback
from openpyxl import load_workbook, Workbook
from Chat_ProcessSingleInput import process_question
from L5XGen_Routine import ProcessRoutineExcel, GenerateRoutine
from L5XGen_AOI import GenerateAOI

def process_excel_file(input_file_path: str, mode: str, log_file_path: str = "LogExcel.xlsx", output_l5x_path: str = "Output.L5X"):
    """
    Processes an Excel file and generates Routine or AOI L5X file based on mode.
    Returns True on success, False on failure.
    """
    try:
        # Load input Excel
        try:
            wb = load_workbook(input_file_path)
            ws = wb.active
        except Exception as e:
            raise ValueError(f"❌ Failed to read Excel file: {e}")

        # Prepare log workbook
        try:
            if os.path.exists(log_file_path):
                os.remove(log_file_path)
        except Exception as e:
            raise IOError(f"❌ Failed to delete old log file: {e}")

        log_wb = Workbook()
        log_ws = log_wb.active
        log_ws.title = "Log"
        log_ws.append([
            "LineText", "Model Inferred Keywords", "Model Inferred Operation",
            "Model Output", "Detected Keywords", "Detected Instruction", "Match"
        ])

        num = 0

        # Process each question row
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            question_cell = row[0]
            ilcode_cell = row[1]
            response_time_cell = row[2]

            question = question_cell.value
            if not question:
                continue

            num += 1
            print(f"\n[{num}]---------------------------------------------------------------------------------------------")
            start_time = time.time()

            try:
                result = process_question(question)

                ilcode_cell.value = result["final_output"]
                response_time_cell.value = result["time_taken"]

                validated = result["validated_instructions"]
                for i, output in enumerate(validated["ops"]):
                    log_ws.append([
                        question,
                        validated["model_keywords_list"][i],
                        validated["inferred_operations"][i],
                        output,
                        ", ".join(validated["found_keywords_list"][i]) or "None",
                        f"{validated['detected_instrs'][i]}({validated['operand_counts'][i]})"
                        if validated["detected_instrs"][i] else "None",
                        validated["matches"][i]
                    ])

            except Exception as e:
                print(f"❌ Error processing question in row {row_idx}: {e}")
                traceback.print_exc()
                ilcode_cell.value = "Error"
                response_time_cell.value = "Error"

        # Save updated input file
        wb.save(input_file_path)

        # Generate L5X text
        try:
            text_output = ProcessRoutineExcel(input_file_path)
        except Exception as e:
            print(f"❌ Error extracting text output: {e}")
            traceback.print_exc()
            return False  # Failed to extract text

        # Generate L5X file
        if mode == 'routine':
            routine_result = GenerateRoutine(text_output, output_file=output_l5x_path)
            if routine_result["success"]:
                print(f"✅ Routine file generated: {output_l5x_path}")
            else:
                print(f"❌ Failed to generate routine: {routine_result['error']}")
                return False
        elif mode == 'aoi':
            aoi_result = GenerateAOI(text_output, output_file=output_l5x_path)
            if aoi_result["success"]:
                print(f"✅ AOI file generated: {output_l5x_path}")
            else:
                print(f"❌ Failed to generate AOI: {aoi_result['error']}")
                return False
        else:
            print(f"⚠️ Unknown mode: {mode}. No L5X file generated.")
            return False

        # Save log file
        try:
            log_wb.save(log_file_path)
            print(f"✅ Log file generated: {log_file_path}")
        except Exception as e:
            print(f"❌ Error saving log file: {e}")
            traceback.print_exc()
            return False

    except Exception as e:
        print(f"❌ Unhandled processing error: {e}")
        traceback.print_exc()
        return False

    return True  # ✅ SUCCESS
